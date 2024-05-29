from ragtime.base.data_type import (
    RagtimeList,
    QA,
    Answers,
    Answer,
    UpdateTypes,
    Fact,
    Facts,
)

from ragtime.config import (
    DEFAULT_FACTS_COL,
    DEFAULT_HUMAN_EVAL_COL,
    DEFAULT_HEADER_SIZE,
    DEFAULT_HUMAN_EVAL_COL,
    DEFAULT_ANSWERS_COL,
    DEFAULT_QUESTION_COL,
    DEFAULT_WORKSHEET,
    logger,
)

from collections import defaultdict

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from typing import Optional
from pydantic import Field
from tabulate import tabulate
from pathlib import Path
import json


class Expe(RagtimeList[QA]):
    meta: Optional[dict] = {}
    # json_path: Path = Field(None, exclude=True)
    folder: Path = Field(None, exclude=True)
    file_name: str = Field(None, exclude=True)

    @property
    def path(self) -> Path:
        return self.folder / Path(self.file_name)

    def __init__(self, folder: Path = None, file_name: str = None):
        super().__init__()
        if not (folder and file_name):
            return
        self.folder = folder
        self.file_name = file_name
        self._load_from_json(path=self.path)

    def _load_from_json(self, path: Path):
        with open(path, mode="r", encoding="utf-8") as file:
            data: list = json.load(file)
            qa_list: dict = data
            if "meta" in data:
                self.meta = data["meta"]
                qa_list = data["items"]
            for json_qa in qa_list:
                qa: QA = QA(**json_qa)
                self.append(qa)

    def stats(self) -> dict:
        """
        Returns stats about the expe :
            - number of models,
            - number of questions,
            - number of facts,
            - number of answers,
            - number of human eval,
            - number of auto eval
        """
        res: dict = {}
        res["questions"] = len([qa for qa in self if qa.question.text])
        res["chunks"] = len([c for qa in self for c in qa.chunks if c])
        res["facts"] = len([f for qa in self for f in qa.facts if f])
        res["models"] = len(self[0].answers)
        res["answers"] = len([a for qa in self for a in qa.answers if a.text])
        res["human eval"] = len(
            [a for qa in self for a in qa.answers if a.eval and a.eval.human]
        )
        res["auto eval"] = len(
            [a for qa in self for a in qa.answers if a.eval and a.eval.auto]
        )
        return res

    def filter_answer(self, llm_facts_name: str):
        """
        Filters the current Expe object to include only the QA pairs with answers from the specified LLM.
        """
        for qa in self:
            filtered_answers = [
                a
                for a in qa.answers
                if a.llm_answer and a.llm_answer.name == llm_facts_name
            ]
            qa.answers = Answers(items=filtered_answers)


def analyse_expe_folder(path: Path):
    if not path.is_dir():
        raise Exception(f'"{path}" is not a folder - please provide one')
    print(f'In "{path}":')
    res: defaultdict = defaultdict(list)
    for f in [f for f in path.iterdir() if f.is_file() and f.suffix == ".json"]:
        exp: Expe = Expe(json_path=f)
        res["File"].append(f.name)
        for k, v in exp.stats():
            res[k].append(v)

    print(tabulate(res, headers="keys"))


def update_from_spreadsheet(
    expe: Expe,
    path: Path,
    update_type: UpdateTypes,
    data_col: int = None,
    question_col: int = DEFAULT_QUESTION_COL - 1,
    answer_col: int = DEFAULT_ANSWERS_COL - 1,
    sheet_name: str = DEFAULT_WORKSHEET,
    header_size: int = DEFAULT_HEADER_SIZE,
):
    """
    Updates data from a spreadsheet, e.g. human evaluation or facts
    Args:
        - data_col (int): indicates the column number (starts at 0) from where the data will be imported in the spreadsheet if None (default), default column values are used, i.e.
        - DEFAULT_FACTS_COL if update_type==Facts and
        - DEFAULT_HUMAN_EVAL_COL if update_type==human_eval
        - update_type (UpdateTypes): can be "human_eval" or "facts"
        - question_col (int): indicates the column number (starts at 0) where the questions are - default: DEFAULT_QUESTION_COL-1 (0 based)
        - answer_col (in): used if update_type==human_eval, since the eval entered in the spreadsheet has to be matched with a specific answer
    """

    def starts_with_num(fact: str) -> bool:
        result: bool = False
        if "." in fact:
            try:
                dummy: int = int(fact[: fact.find(".")])
                result = True
            except (TypeError, ValueError):
                pass
        return result

    wb: Workbook = load_workbook(path)
    ws: Worksheet = wb[sheet_name]
    cur_qa: QA = None
    if not data_col:
        data_cols: dict = {
            UpdateTypes.facts: DEFAULT_FACTS_COL,
            UpdateTypes.human_eval: DEFAULT_HUMAN_EVAL_COL,
        }
        data_col = data_cols[update_type] - 1

    # the new facts to replace the old ones in the current QA
    new_facts: Facts = Facts()

    # For each row in the worksheet
    for i, row in enumerate(ws.iter_rows(min_row=header_size + 1), start=1):
        # a question is in the current row, so a new question starts
        if row[question_col].value:
            if cur_qa:  # not first question
                cur_qa.facts = new_facts
            # get the corresponding QA in the Expe
            cur_qa = next(
                (
                    qa
                    for qa in expe
                    if qa.question.text.lower() == row[question_col].value.lower()
                ),
                None,
            )
            new_facts: Facts = Facts()

        if not cur_qa:
            continue
        # QA and question in the worksheet is made
        data_in_ws = row[data_col].value
        if not data_in_ws:
            continue

        if update_type == UpdateTypes.facts:  # Update FACTS
            # if the fact in the ws does not start with a number, add it
            if not starts_with_num(data_in_ws):
                data_in_ws = f"{len(new_facts) + 1}. {data_in_ws}"
            new_facts.append(Fact(text=data_in_ws))
        elif update_type == UpdateTypes.human_eval:  # Update HUMAN EVAL
            answer_text: str = row[answer_col].value
            cur_ans: Answer = next(
                (a for a in cur_qa.answers if a.text == answer_text), None
            )
            if not cur_ans:
                logger.warn(
                    f'Cannot find Answer corresponding with the human eval "{data_in_ws}" - Answer should contain the text "{answer_text}"'
                )
            # corresponding Answer has been found
            try:
                human_eval: int = int(data_in_ws)
                cur_ans.eval.human = human_eval
            except (TypeError, ValueError):
                logger.warn(
                    f'Human eval should be a value between 0 and 1 - cannot use "{data_in_ws}" as found in line {i}'
                )
