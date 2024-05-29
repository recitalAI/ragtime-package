from ragtime.base.exporter import Exporter, Expe, Path


from ragtime.config import (
    DEFAULT_HEADER_SIZE,
    DEFAULT_SPREADSHEET_TEMPLATE,
    DEFAULT_WORKSHEET,
    RagtimeException,
)

from ragtime.base.data_type import QA

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from datetime import datetime
from pathlib import Path
from copy import copy
import shutil
import re


class Spreadsheet(Exporter):
    sheet_name: str = DEFAULT_WORKSHEET
    header_size: int = DEFAULT_HEADER_SIZE
    template_path: Path = DEFAULT_SPREADSHEET_TEMPLATE
    _force_extension: str = ".xlsx"

    """
    Saves Expe to a spreadsheet - can generate a suffix for the filename
    Returns the Path of the file actually saved
    """

    def save(self, expe: Expe, folder: Path = None, file_name: str = None) -> Path:
        path: Path = self._file_check_before_writing(expe, folder, file_name)
        # Prepare the result file
        shutil.copy(self.template_path, path)  # Copy template
        wb = load_workbook(path)
        wb.iso_dates = True
        ws = wb[self.sheet_name]  # Create worksheet

        # Retrieve sst configuration from original file
        # ws_conf is a list of str, each element describes the path of the data to be added in the current row
        ws_conf = [cell.value for cell in ws[self.header_size + 1]]

        # Write the values at specific rows - they are defined in second row, below the one describing the value to insert
        for cell in ws[self.header_size + 2]:
            # read the row just after the conf row - it contains configuration for specific rows
            # if a value is present, analyse it - it should contain a "row" indication e.g. "answers[0].full_name, row=1"
            # special token # used to indicate question number
            if not cell.value or cell.value == "#":
                continue

            val: str = cell.value
            row: int = int(val[val.find("row=") + len("row=") :])
            if row < 1:
                raise RagtimeException(
                    f'The row value "row={row}" specified in cell {cell.coordinate} is invalid and must be greater than 0'
                )
            # write the value since it does not need to be done for each row
            p: str = val[: val.find(",")]
            # get the first non empty value in the required column
            val = next((qa.get_attr(p) for qa in self if qa.get_attr(p)), "")
            ws.cell(row=row, column=cell.column, value=val)

        qa: QA
        row: int = self.header_size + 1
        col_with_formulas: dict[int, str] = {
            c: ws.cell(column=c, row=row).value
            for c in range(1, ws.max_column)
            if ws.cell(column=c, row=row).value
            and str(ws.cell(column=c, row=row).value)[0] == "="
        }

        def forEachQA(row, question_number):
            def isFormula(value):
                return value[0] == "#"

            def onTokenOrPath(value, num_q):
                # special token # used to indicate question number
                if value == "#":
                    return [num_q]
                # else it's a path to get a value in QA
                # and write a blank if nothing is found
                value = qa.get_attr(p)
                return [""] if value is None or value == [] else value

            def WriteValues(values, column, row, header_size):
                def standard_conversion_to_string(value):
                    # Do standard conversions to string
                    if isinstance(value, list):
                        value = str(value)
                    if isinstance(value, datetime):
                        value = value.strftime("%d/%m/%Y %H:%M:%S")
                    return value

                def remove_illegal_characters(value):
                    # Remove illegal characters
                    return ILLEGAL_CHARACTERS_RE.sub("", str(value))

                def write_value(value, column, row, header_size):
                    # Write value
                    ws.cell(row=row, column=column).value = value
                    # From second row copy cell style from the one up
                    if row > header_size:
                        header_style = ws.cell(row=header_size, column=column)._style
                        ws.cell(row=row, column=column)._style = copy(header_style)

                for offset, value in enumerate(values):
                    value = standard_conversion_to_string(value)
                    value = remove_illegal_characters(value)
                    write_value(
                        column=column,
                        row=row + offset,
                        header_size=header_size + 1,
                        value=value,
                    )
                return row + offset + 1

            next_row: int = 0
            for column, p in enumerate(ws_conf, start=1):
                if isFormula(p):
                    continue
                value = onTokenOrPath(p, question_number)
                value = value if isinstance(value, list) else [value]
                last_row = WriteValues(value, column, row, self.header_size)
                next_row = max(next_row, last_row)
            return next_row

        for num_q, qa in enumerate(expe, start=1):  # write each row in expe
            row = forEachQA(row, num_q)

        # extend the formulas
        for row in range(self.header_size + 1, ws.max_row):
            for col, formula in col_with_formulas.items():
                # simply adjust the row number in the formula
                cell_refs: set = set(re.findall(r"[A-Z]+[0-9]+", formula))
                for cell_ref in cell_refs:
                    new_cell_ref: str = cell_ref.replace(
                        str(self.header_size + 1), str(row)
                    )
                    formula = formula.replace(cell_ref, new_cell_ref)
                ws.cell(row=row, column=col, value=formula)

        wb.save(path)
        return path
